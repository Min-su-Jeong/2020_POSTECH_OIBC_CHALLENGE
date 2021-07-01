import requests
import json
from openpyxl import Workbook

# INPUT BY YOURSELF
apikey = 'Input your own apikey'
date = '2020-10-14'

# WORKING ON
req_url = 'https://competition.api.henergy.xyz/api/recent_pv'
request_datas = {'api_key': apikey, 'date':date}

returned_data = requests.post(req_url, data=json.dumps(request_datas))
pv_data = json.loads(returned_data.text)

# OUTPUT
data=list(pv_data.values())
data.pop()
answer = sum(data, [])
answer.reverse()

# Write and Save Excel
wb = Workbook()

sheet = wb.active
file_name = 'data_new.xlsx'
sheet_title = 'data'
sheet.column_dimensions['A'].width = 25
sheet.column_dimensions['B'].width = 10
sheet.append(['datetime','energy'])

          
for i in range(len(answer)):
    sheet.cell(row=i+2, column=1).value=answer[i][0]
    sheet.cell(row=i+2, column=2).value=answer[i][1]

wb.save('data_new.xlsx')

print("Complete Saved!")
